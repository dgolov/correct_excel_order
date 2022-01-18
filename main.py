from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException

import json
import sys


def open_excel(path: str, sheet_name: str) -> (tuple, None):
    """ Open excel order
    :param path: path in arg 1
    :param sheet_name: sheet name where to look
    :return:
    """
    print('Loading document...')
    try:
        wb = load_workbook(path)
    except InvalidFileException:
        print(f'No such file or directory: {path}')
        return None
    try:
        sheet = wb.get_sheet_by_name(sheet_name)
        return wb, sheet
    except KeyError:
        print('Incorrect sheet name')
        return None


def correcting_file(wb, sheet, coll_key: int, coll_value: int, data: json) -> None:
    """ Read end correct order
    :param wb: work book object
    :param sheet: sheet object
    :param coll_key: column number to search by key
    :param coll_value: column number for replacement
    :param data: correct data
    :return:
    """
    print('Parsing document...')
    is_correcting = False
    for row in range(1, sheet.max_row):
        result_to_replace = data.get(sheet.cell(row=row, column=coll_key).value)
        if result_to_replace:
            is_correcting = True
            sheet.cell(row=row, column=coll_value).value = result_to_replace

    if not is_correcting:
        print('No changes have been made')
    else:
        wb.save('./correct_order.xlsx')
    return


def load_json(path: str) -> (json, None):
    """ load correct data in json file
    :param path: path in arg 2
    :return: json data
    """
    print('Open json file')
    try:
        with open(path, 'r') as file:
            data = json.load(file)
            return data
    except FileNotFoundError:
        print(f'No such file or directory: {path}')
        return None


def make_run(path_to_excel: str, path_to_json: str) -> int:
    """ Entry point to correct
    :param path_to_excel: path in arg 1
    :param path_to_json:  path in arg 3
    :return: result status
    """
    data = load_json(path_to_json)
    wb = Workbook()
    sheet = wb.create_sheet('Sheet')

    row = 0
    print('Write json to Excel')
    for key, value in data.items():
        row += 1
        sheet.cell(row=row, column=1).value = key
        sheet.cell(row=row, column=2).value = value

    wb.save(path_to_excel)
    return 1


def run_correct(path_to_excel: str, path_to_json: str, sheet: str) -> int:
    """ Entry point to correct
    :param path_to_excel: path in arg 1
    :param sheet:  path in arg 2
    :param path_to_json:  path in arg 3
    :return: result status
    """
    key_number = int(input('Input the key coll number: '))
    value_number = int(input('Input the replacement coll number: '))

    correct_data = load_json(path_to_json)
    if not correct_data:
        return 0

    try:
        order = open_excel(
            path=path_to_excel,
            sheet_name=sheet
        )
        correcting_file(*order, coll_key=key_number, coll_value=value_number, data=correct_data)
        return 1
    except TypeError:
        return 0


if __name__ == '__main__':
    if len(sys.argv) == 2 and (sys.argv[1] == '-h' or sys.argv[1] == '--help'):
        print('python main.py <path to excel> <sheet name> <path to json>')
        result = 0
    elif len(sys.argv) < 4:
        print('Missing number of arguments passed (requires 3)')
        result = 0
    elif len(sys.argv) == 4 and (sys.argv[2] == '-m' or sys.argv[2] == '--make'):
        result = make_run(path_to_excel=sys.argv[1], path_to_json=sys.argv[3])
    else:
        result = run_correct(path_to_excel=sys.argv[1], sheet=sys.argv[2], path_to_json=sys.argv[3])
    print('Success!') if result else print('Fail')
